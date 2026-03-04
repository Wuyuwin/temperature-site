import os
from datetime import datetime, timezone, timedelta
from functools import wraps
from flask import Flask, render_template, request, Response, redirect, url_for
from sqlalchemy import create_engine, text, event
from dotenv import load_dotenv
from flask import send_file

try:
    from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
except Exception:
    ZoneInfo = None
    ZoneInfoNotFoundError = Exception

load_dotenv()
app = Flask(__name__, instance_relative_config=True)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret")

APP_TZ = os.getenv("APP_TZ", "Asia/Taipei")

# Database
db_url = os.getenv("DATABASE_URL", "sqlite:///instance/temperature.db")
os.makedirs(app.instance_path, exist_ok=True)
if db_url.startswith("sqlite:///") and "instance/" in db_url:
    db_url = "sqlite:///" + os.path.join(app.instance_path, "temperature.db")
engine = create_engine(db_url, future=True, echo=False)

@event.listens_for(engine, "connect")
def _set_sqlite_pragmas(dbapi_connection, connection_record):
    cursor = dbapi_connection.cursor()
    cursor.execute("PRAGMA journal_mode=WAL;")
    cursor.execute("PRAGMA synchronous=NORMAL;")
    cursor.execute("PRAGMA busy_timeout=10000;")
    cursor.execute("PRAGMA temp_store=MEMORY;")
    cursor.close()

# Create tables
with engine.begin() as conn:
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            person_name TEXT NOT NULL,
            location TEXT NOT NULL,
            working_time TEXT NOT NULL,
            symptom TEXT,
            note TEXT,
            temperature_c REAL NOT NULL,
            timestamp_utc TEXT NOT NULL
        )
    """))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_timestamp ON records(timestamp_utc);"))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_person_ts ON records(person_name, timestamp_utc);"))

    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            active INTEGER NOT NULL DEFAULT 1,
            created_utc TEXT NOT NULL
        )
    """))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_emp_active ON employees(active);"))

# Basic Auth
ADMIN_USER = os.getenv("BASIC_AUTH_USER")
ADMIN_PASS = os.getenv("BASIC_AUTH_PASS")

def requires_auth(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.authorization
        if not auth or auth.username != ADMIN_USER or auth.password != ADMIN_PASS:
            return Response("需要後台帳密\n", 401, {"WWW-Authenticate": 'Basic realm="Admin Area"'})
        return f(*args, **kwargs)
    return wrapper

def _date_range_utc(selected_date=None):
    if ZoneInfo is None:
        if selected_date:
            day = datetime.strptime(selected_date, "%Y-%m-%d").date()
        else:
            day = datetime.now(timezone.utc).date()

        start = datetime(day.year, day.month, day.day, tzinfo=timezone.utc)
        end = start + timedelta(days=1)
        return start, end, day.isoformat()

    try:
        tz = ZoneInfo(APP_TZ)
    except ZoneInfoNotFoundError:
        if selected_date:
            day = datetime.strptime(selected_date, "%Y-%m-%d").date()
        else:
            day = datetime.now(timezone.utc).date()

        start = datetime(day.year, day.month, day.day, tzinfo=timezone.utc)
        end = start + timedelta(days=1)
        return start, end, day.isoformat()

    # timezone version
    if selected_date:
        day = datetime.strptime(selected_date, "%Y-%m-%d").date()
        start_local = datetime(day.year, day.month, day.day, tzinfo=tz)
    else:
        now_local = datetime.now(tz)
        start_local = datetime(now_local.year, now_local.month, now_local.day, tzinfo=tz)

    end_local = start_local + timedelta(days=1)

    start_utc = start_local.astimezone(timezone.utc)
    end_utc = end_local.astimezone(timezone.utc)

    return start_utc, end_utc, start_local.date().isoformat()

@app.get("/")
def index():
    return render_template("index.html", msg=request.args.get("msg"), error=request.args.get("error"))

@app.get("/qr.png")
def qr_png():
    import qrcode
    from io import BytesIO
    # 產生「目前網頁的 base URL」做 QR
    target = request.url_root
    img = qrcode.make(target)
    bio = BytesIO()
    img.save(bio, format="PNG")
    bio.seek(0)
    return send_file(bio, mimetype="image/png")

@app.post("/submit")
def submit():
    person_name = (request.form.get("person_name") or "").strip()
    location = (request.form.get("location") or "").strip()
    working_time = (request.form.get("working_time") or "").strip()
    temperature_c = request.form.get("temperature_c")
    note = request.form.get("note") or None
    symptoms = request.form.getlist("symptom")
    symptom_str = ",".join(symptoms) if symptoms else "0"

    if not person_name or not location or not working_time or not temperature_c:
        return redirect(url_for("index", error="姓名、科別、上班時間、體溫 為必填"))
    try:
        t = float(temperature_c)
        if not (34.0 <= t <= 42.0):
            raise ValueError
    except Exception:
        return redirect(url_for("index", error="體溫格式錯誤或超出範圍（34.0~42.0℃）"))

    now = datetime.now(timezone.utc).isoformat()
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO records (person_name, location, working_time, symptom, note, temperature_c, timestamp_utc)
            VALUES (:n, :l, :w, :s, :no, :t, :ts)
        """), dict(n=person_name, l=location, w=working_time, s=symptom_str, no=note, t=t, ts=now))
    return redirect(url_for("index", msg="已送出，謝謝！"))

def _load_employees():
    with engine.begin() as conn:
        employees = conn.execute(text("""
            SELECT id, name, active, created_utc
            FROM employees
            ORDER BY active DESC, name ASC, id ASC
        """)).mappings().all()
        active_count = conn.execute(text("SELECT COUNT(*) FROM employees WHERE active=1")).scalar_one()
    return employees, active_count

def _load_date_audit(selected_date=None):
    start_utc, end_utc, date_label = _date_range_utc(selected_date)
    start_s = start_utc.isoformat()
    end_s = end_utc.isoformat()

    with engine.begin() as conn:
        last_today = conn.execute(text("""
            SELECT rr.person_name, rr.location, rr.working_time, rr.symptom, rr.note, rr.temperature_c, rr.timestamp_utc
            FROM records rr
            JOIN (
              SELECT person_name, MAX(timestamp_utc) AS max_ts
              FROM records
              WHERE timestamp_utc >= :start AND timestamp_utc < :end
              GROUP BY person_name
            ) m ON rr.person_name=m.person_name AND rr.timestamp_utc=m.max_ts
            WHERE rr.timestamp_utc >= :start AND rr.timestamp_utc < :end
        """), dict(start=start_s, end=end_s)).mappings().all()

        last_map = {r["person_name"]: r for r in last_today}

        employees = conn.execute(text("""
            SELECT id, name
            FROM employees
            WHERE active=1
            ORDER BY name ASC
        """)).mappings().all()

    audit_rows = []
    for e in employees:
        r = last_map.get(e["name"])
        if r:
            audit_rows.append({
                "name": e["name"],
                "status": "已填",
                "location": r.get("location"),
                "working_time": r.get("working_time"),
                "symptom": r.get("symptom"),
                "note": r.get("note"),
                "temperature_c": r.get("temperature_c"),
                "timestamp_utc": r.get("timestamp_utc"),
            })
        else:
            audit_rows.append({
                "name": e["name"],
                "status": "未填",
                "location": "",
                "working_time": "",
                "symptom": "",
                "note": "",
                "temperature_c": None,
                "timestamp_utc": "",
            })
    filled_count = sum(1 for x in audit_rows if x["status"] == "已填")
    missing_count = len(audit_rows) - filled_count
    audit_rows.sort(
        key=lambda x: (
            x["status"] == "已填",
            x["timestamp_utc"] if x["timestamp_utc"] else "",
        ),
        reverse=True
    )
    return audit_rows, filled_count, missing_count, date_label

@app.get("/admin")
@requires_auth
def admin_home():
    selected_date = (request.args.get("date") or "").strip() or None
    try:
        per_page = int(request.args.get("per_page", 10))
    except Exception:
        per_page = 10

    if per_page not in (10, 20, 30):
        per_page = 10

    try:
        page = int(request.args.get("page", 1))
    except Exception:
        page = 1

    page = max(page, 1)
    offset = (page - 1) * per_page

    employees, active_count = _load_employees()
    audit_rows, filled_count, missing_count, date_label = _load_date_audit(selected_date)

    total = len(audit_rows)
    total_pages = (total + per_page - 1) // per_page if total else 1
    audit_page = audit_rows[offset:offset + per_page]

    return render_template(
        "admin.html",
        employees=employees,
        employees_active_count=active_count,
        audit_rows=audit_rows,
        audit_page=audit_page,
        filled_count=filled_count,
        missing_count=missing_count,
        date_label=date_label,
        selected_date=(selected_date or date_label),
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        msg=request.args.get("msg"),
        error=request.args.get("error")
    )

@app.post("/admin/employees/import")
@requires_auth
def employees_import():
    f = request.files.get("file")
    if not f or not f.filename.lower().endswith(".xlsx"):
        return redirect(url_for("admin_home", error="請上傳 .xlsx 檔案"))

    try:
        from openpyxl import load_workbook
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [(str(c).strip() if c is not None else "") for c in header_row]

        if "員工名字" not in headers:
            return redirect(url_for("admin_home", error="Excel 第一列需包含欄位：員工名字"))

        idx = headers.index("員工名字")
        names = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[idx] if row and idx < len(row) else None
            if val is None:
                continue
            name = str(val).strip()
            if name:
                names.append(name)

        if not names:
            return redirect(url_for("admin_home", error="沒有讀到任何員工名字"))

        inserted = 0
        skipped = 0
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM employees"))
            conn.execute(text("DELETE FROM sqlite_sequence WHERE name='employees'"))
            for name in names:
                conn.execute(text("""
                    INSERT INTO employees(name, active, created_utc)
                    VALUES (:n, 1, :ts)
                """), dict(
                    n=name,
                    ts=datetime.now(timezone.utc).isoformat()
                ))
                inserted += 1

        return redirect(url_for("admin_home",msg=f"匯入完成：已取代全部員工資料，共 {inserted} 筆"))
    except Exception as e:
        return redirect(url_for("admin_home", error="匯入失敗：" + str(e)[:200]))

@app.get("/admin/audit.xlsx")
@requires_auth
def export_audit_xlsx():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import datetime

    selected_date = (request.args.get("date") or "").strip() or None
    audit_rows, filled_count, missing_count, date_label = _load_date_audit(selected_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "稽核"

    headers = ["姓名", "狀態", "科別", "上班時間", "症狀", "備註", "體溫(℃)", "時間(UTC)"]
    ws.append(headers)

    for r in audit_rows:
        ws.append([
            r["name"],
            r["status"],
            r.get("location",""),
            r.get("working_time",""),
            r.get("symptom",""),
            r.get("note",""),
            r["temperature_c"] if r["temperature_c"] is not None else "",
            r.get("timestamp_utc",""),
        ])

    for i, col in enumerate(ws.columns, start=1):
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = min(max(length + 2, 10), 48)

    bio = BytesIO()
    wb.save(bio); bio.seek(0)
    filename = f"audit_{date_label}.xlsx"
    return Response(
        bio.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/healthz")
def healthz():
    return {"status": "ok"}

if __name__ == "__main__":
    from waitress import serve
    serve(app, host="0.0.0.0", port=8000, threads=8)
